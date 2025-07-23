import csv
import zipfile
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from script_os1 import ZIP_DIR


def test_pdf_file():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open("example.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()
            assert "Pocket Money" in text

def test_xlsx():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open("example.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            value = sheet.cell(row=9, column=1).value
            name = "February"
            assert name in value

def test_csv():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open("example.csv") as csv_file:
            content = csv_file.read().decode("utf-8-sig")
            csvreader = list(csv.reader(content.splitlines()))
            fourth_row = csvreader[4]
            assert fourth_row[0] == "Pocket Money"








